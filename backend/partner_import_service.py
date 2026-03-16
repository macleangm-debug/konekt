"""
Partner Import Service
Parse CSV and Excel files for bulk listing imports
"""
import csv
import io
from typing import List, Dict


EXPECTED_COLUMNS = [
    "listing_type",
    "product_family",
    "service_family",
    "sku",
    "slug",
    "name",
    "short_description",
    "description",
    "category",
    "subcategory",
    "brand",
    "country_code",
    "regions",
    "currency",
    "base_partner_price",
    "partner_available_qty",
    "partner_status",
    "lead_time_days",
    "images",
    "documents",
]


def normalize_row(row: dict) -> dict:
    """Normalize and clean a row from import file"""
    clean = {}

    for key in EXPECTED_COLUMNS:
        clean[key] = row.get(key)

    # String fields
    clean["listing_type"] = (str(clean.get("listing_type") or "product")).strip().lower()
    clean["product_family"] = (str(clean.get("product_family") or "")).strip() or None
    clean["service_family"] = (str(clean.get("service_family") or "")).strip() or None
    clean["sku"] = (str(clean.get("sku") or "")).strip()
    clean["slug"] = (str(clean.get("slug") or "")).strip()
    clean["name"] = (str(clean.get("name") or "")).strip()
    clean["short_description"] = (str(clean.get("short_description") or "")).strip()
    clean["description"] = (str(clean.get("description") or "")).strip()
    clean["category"] = (str(clean.get("category") or "")).strip()
    clean["subcategory"] = (str(clean.get("subcategory") or "")).strip()
    clean["brand"] = (str(clean.get("brand") or "")).strip()
    clean["country_code"] = (str(clean.get("country_code") or "")).strip().upper()
    clean["currency"] = (str(clean.get("currency") or "TZS")).strip().upper()
    clean["partner_status"] = (str(clean.get("partner_status") or "in_stock")).strip().lower()

    # List fields (comma-separated)
    clean["regions"] = [
        x.strip() for x in str(clean.get("regions") or "").split(",") if x.strip()
    ]
    clean["images"] = [
        x.strip() for x in str(clean.get("images") or "").split(",") if x.strip()
    ]
    clean["documents"] = [
        x.strip() for x in str(clean.get("documents") or "").split(",") if x.strip()
    ]

    # Numeric fields
    try:
        clean["base_partner_price"] = float(clean.get("base_partner_price") or 0)
    except (ValueError, TypeError):
        clean["base_partner_price"] = 0

    try:
        clean["partner_available_qty"] = float(clean.get("partner_available_qty") or 0)
    except (ValueError, TypeError):
        clean["partner_available_qty"] = 0

    try:
        clean["lead_time_days"] = int(float(clean.get("lead_time_days") or 2))
    except (ValueError, TypeError):
        clean["lead_time_days"] = 2

    return clean


def parse_csv_bytes(file_bytes: bytes) -> List[Dict]:
    """Parse CSV file bytes into list of dictionaries"""
    # Try different encodings
    for encoding in ["utf-8-sig", "utf-8", "latin-1"]:
        try:
            decoded = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode file. Please use UTF-8 encoding.")

    reader = csv.DictReader(io.StringIO(decoded))
    rows = [normalize_row(row) for row in reader]
    return rows


def parse_xlsx_bytes(file_bytes: bytes) -> List[Dict]:
    """Parse Excel file bytes into list of dictionaries"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Excel support requires openpyxl. Please use CSV format.")

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is headers
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    output = []

    for row in rows[1:]:
        obj = {}
        for idx, header in enumerate(headers):
            obj[header] = row[idx] if idx < len(row) else None
        output.append(normalize_row(obj))

    return output


def parse_partner_import(file_name: str, file_bytes: bytes) -> List[Dict]:
    """Parse import file based on extension"""
    lower = (file_name or "").lower()

    if lower.endswith(".csv"):
        return parse_csv_bytes(file_bytes)
    elif lower.endswith(".xlsx"):
        return parse_xlsx_bytes(file_bytes)
    else:
        raise ValueError("Only CSV and XLSX files are supported")


def get_template_headers() -> List[str]:
    """Return expected column headers"""
    return EXPECTED_COLUMNS.copy()
